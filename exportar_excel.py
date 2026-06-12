"""Exportar reportes a Excel (.xlsx)."""

from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta
from pathlib import Path

from config import DATA_DIR
from persistencia import reporte_dia, listar_ventas, conexion as _conexion


def _fmt(n: int | None) -> str:
    """Formatea entero como precio colombiano para celdas de texto."""
    return f"${int(n or 0):,}".replace(",", ".")


def exportar_ventas_excel(fecha: str | None = None, ruta: Path | None = None, output=None):
    """
    Genera Excel con 3 hojas para el dia indicado (o hoy si fecha=None):
      1. Resumen del dia   — cards de totales + tabla de turnos
      2. Ventas detalladas — una fila por factura
      3. Productos vendidos — una fila por linea de venta
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError as e:
        raise ImportError("Instala openpyxl: pip install openpyxl") from e

    if fecha is None:
        fecha = datetime.now().strftime("%Y-%m-%d")

    datos = reporte_dia(fecha)
    r     = datos["resumen"]
    turnos_data = datos["turnos"]
    ventas_data = datos["ventas"]

    # Para items de cada venta necesitamos los detalles completos
    from persistencia import obtener_venta_por_id

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    destino = ruta or (DATA_DIR / f"alvarez_{fecha}.xlsx")

    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────────────────────
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    h_fill  = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center")

    gold_fill  = PatternFill(start_color="F0A500", end_color="F0A500", fill_type="solid")
    gold_font  = Font(bold=True, size=13)
    green_fill = PatternFill(start_color="0f2010", end_color="0f2010", fill_type="solid")
    green_font = Font(bold=True, color="4ade80")
    red_fill   = PatternFill(start_color="1e0a0a", end_color="1e0a0a", fill_type="solid")
    red_font   = Font(bold=True, color="f87171")

    thin = Side(style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header(ws, row, cols):
        for col_idx, title in enumerate(cols, 1):
            c = ws.cell(row=row, column=col_idx, value=title)
            c.font  = h_font
            c.fill  = h_fill
            c.alignment = h_align

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 1 — Resumen del dia
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen del dia"

    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value    = f"Alvarez Fast Food — {fecha}"
    title_cell.font     = Font(bold=True, size=14, color="F0A500")
    title_cell.alignment = Alignment(horizontal="center")

    # Cards de resumen
    labels = [
        ("Total del dia",       r["total_ventas"]),
        ("Efectivo",            r["total_efectivo"]),
        ("Nequi",               r["total_nequi"]),
        ("Transferencia/Otros", r["total_transferencia"] + r["total_otros"]),
        ("Vueltos dados",       r["total_vueltos"]),
        ("Total facturas",      r["total_facturas"]),
    ]
    ws1.cell(row=3, column=1, value="RESUMEN").font = Font(bold=True, size=11, color="888888")
    for i, (lbl, val) in enumerate(labels, 4):
        ws1.cell(row=i, column=1, value=lbl).font = Font(color="AAAAAA")
        c = ws1.cell(row=i, column=2, value=val)
        c.number_format = "#,##0"
        c.font = Font(bold=True, color="F0A500", size=12)

    # Tabla de turnos
    row_t = 4 + len(labels) + 2
    ws1.cell(row=row_t - 1, column=1, value="TURNOS DEL DIA").font = Font(bold=True, size=11, color="888888")
    turno_headers = ["Cajero", "Apertura", "Cierre", "Base inicial",
                     "Ventas efectivo", "Vueltos", "Efectivo esperado",
                     "Total ventas", "Facturas", "Estado"]
    set_header(ws1, row_t, turno_headers)
    for t in turnos_data:
        row_t += 1
        row_data = [
            t["cajero"],
            t["hora_apertura"],
            t["hora_cierre"] or "abierto",
            t["efectivo_inicial"],
            t["ventas_efectivo"],
            t["total_vueltos"],
            t["efectivo_esperado"],
            t["total_ventas_turno"],
            t["num_facturas"],
            t["estado"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws1.cell(row=row_t, column=col_idx, value=val)
            if col_idx in (4, 5, 6, 7, 8):
                c.number_format = "#,##0"
            if t["estado"] == "abierto":
                c.fill = green_fill
                c.font = green_font
            if col_idx == 7:
                c.font = gold_font

    for col, w in zip("ABCDEFGHIJ", [14, 10, 10, 16, 18, 14, 18, 14, 10, 10]):
        ws1.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 2 — Ventas detalladas
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Ventas detalladas")
    v_headers = ["Factura", "Hora", "Cajero", "Metodo", "Entrega",
                 "Monto recibido", "Vuelto", "Total"]
    set_header(ws2, 1, v_headers)

    for v in ventas_data:
        row_data = [
            v["id_factura"],
            v["hora"],
            v["cajero"],
            v["metodo_pago"],
            v["tipo_entrega"],
            v["monto_recibido"] if v["metodo_pago"] == "Efectivo" else None,
            v["vuelto_dado"]    if v["metodo_pago"] == "Efectivo" else None,
            v["total"],
        ]
        ws2.append(row_data)
        row = ws2.max_row
        for col_idx in (6, 7, 8):
            c = ws2.cell(row=row, column=col_idx)
            if c.value is not None:
                c.number_format = "#,##0"
        # Color badge metodo
        badge_cell = ws2.cell(row=row, column=4)
        if v["metodo_pago"] == "Efectivo":
            badge_cell.font = Font(color="4ade80")
        elif v["metodo_pago"] == "Nequi":
            badge_cell.font = Font(color="f87171")

    # Fila total
    fila_tot = ws2.max_row + 2
    ws2.cell(row=fila_tot, column=7, value="TOTAL:").font = Font(bold=True)
    c_tot = ws2.cell(row=fila_tot, column=8, value=r["total_ventas"])
    c_tot.number_format = "#,##0"
    c_tot.font = gold_font
    c_tot.fill = gold_fill

    for col, w in zip("ABCDEFGH", [18, 10, 14, 16, 12, 16, 14, 14]):
        ws2.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 3 — Productos vendidos
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Productos vendidos")
    p_headers = ["Factura", "Hora", "Cajero", "Producto",
                 "Cantidad", "Precio unit.", "Subtotal", "Notas"]
    set_header(ws3, 1, p_headers)

    ventas_completas = listar_ventas(limite=None) if not ventas_data else None
    # Cargar items por factura del día
    id_set = {v["id_factura"] for v in ventas_data}
    if ventas_completas is None:
        ventas_completas = []
    # Obtener items para cada venta del día
    from persistencia import obtener_venta_por_id as _get
    for v in ventas_data:
        venta_full = _get(v["id_factura"])
        if not venta_full:
            continue
        for item in venta_full.get("items", []):
            ws3.append([
                v["id_factura"],
                v["hora"],
                v["cajero"],
                item.get("producto", item.get("producto_nombre", "")),
                item["cantidad"],
                item["precio_unitario"],
                item["cantidad"] * item["precio_unitario"],
                item.get("notas_modificacion", ""),
            ])
            row = ws3.max_row
            for col_idx in (6, 7):
                ws3.cell(row=row, column=col_idx).number_format = "#,##0"

    for col, w in zip("ABCDEFGH", [18, 10, 14, 36, 10, 14, 14, 24]):
        ws3.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 4 — Insumos del día
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Insumos del dia")
    i_headers = ["Hora", "Insumo", "Cantidad", "Unidad", "Valor Unit.", "Subtotal"]
    set_header(ws4, 1, i_headers)

    total_insumos = 0
    try:
        from persistencia import conexion as _conexion
        with _conexion() as _conn:
            with _conn.cursor() as _cur:
                _cur.execute(
                    """
                    SELECT cd.nombre_insumo, cd.cantidad, cd.unidad,
                           cd.valor_unitario, cd.subtotal, c.fecha_hora, c.notas
                    FROM compras_detalle cd
                    JOIN compras c ON cd.compra_id = c.id
                    WHERE c.fecha = %s
                    ORDER BY c.fecha_hora, cd.id
                    """,
                    (fecha,),
                )
                for row in _cur.fetchall():
                    fh = str(row["fecha_hora"])
                    hora = fh[11:16] if len(fh) > 11 else ""
                    ws4.append([
                        hora,
                        row["nombre_insumo"],
                        float(row["cantidad"]),
                        row["unidad"],
                        int(row["valor_unitario"]),
                        int(row["subtotal"]),
                    ])
                    r4 = ws4.max_row
                    ws4.cell(r4, 5).number_format = "#,##0"
                    ws4.cell(r4, 6).number_format = "#,##0"
                    total_insumos += int(row["subtotal"])
    except Exception:
        pass

    # Filas de resumen
    fila_res = ws4.max_row + 2
    ws4.cell(fila_res, 1, "TOTAL INSUMOS DEL DÍA").font = Font(bold=True)
    c_ti = ws4.cell(fila_res, 6, total_insumos)
    c_ti.number_format = "#,##0"
    c_ti.font = Font(bold=True, color="f87171")

    ws4.cell(fila_res + 1, 1, "TOTAL VENTAS DEL DÍA").font = Font(bold=True)
    c_tv2 = ws4.cell(fila_res + 1, 6, r["total_ventas"])
    c_tv2.number_format = "#,##0"
    c_tv2.font = gold_font

    diferencia = r["total_ventas"] - total_insumos
    ws4.cell(fila_res + 2, 1, "DIFERENCIA").font = Font(bold=True)
    c_dif = ws4.cell(fila_res + 2, 6, diferencia)
    c_dif.number_format = "#,##0"
    c_dif.font = Font(bold=True, color="4ade80" if diferencia >= 0 else "f87171")

    for col, w in zip("ABCDEF", [10, 28, 12, 10, 14, 14]):
        ws4.column_dimensions[col].width = w

    if output is not None:
        wb.save(output)
        return output

    wb.save(destino)
    return destino


# ── UTILIDADES COMPARTIDAS ────────────────────────────────────────────────────

def _estilos_base():
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    h_fill  = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center")
    gold_font  = Font(bold=True, color="F0A500", size=12)
    gold_fill  = PatternFill(start_color="F0A500", end_color="F0A500", fill_type="solid")
    green_font = Font(bold=True, color="4ade80")
    red_font   = Font(bold=True, color="f87171")
    return h_font, h_fill, h_align, gold_font, gold_fill, green_font, red_font


def _set_header(ws, row, cols, h_font, h_fill, h_align):
    for ci, title in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=title)
        c.font = h_font; c.fill = h_fill; c.alignment = h_align


def _sabado_de_semana(d: date) -> date:
    wd = d.weekday()
    if wd == 5: return d
    if wd == 6: return d - timedelta(days=1)
    if wd == 0: return d - timedelta(days=2)
    if wd == 1: return d - timedelta(days=3)
    if wd == 2: return d - timedelta(days=4)
    return d - timedelta(days=wd + 2)


MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


# ════════════════════════════════════════════════════════════════════════════
# FIN DE SEMANA — 4 hojas
# ════════════════════════════════════════════════════════════════════════════

def exportar_fin_semana(fecha: str | None = None, output=None):
    """
    Excel fin de semana (4 hojas):
      1. Resumen semana    — totales ventas / insumos / nómina / ganancia
      2. Ventas del fin de semana — una fila por factura de sáb/dom/lun
      3. Insumos semana    — compras de los 3 días
      4. Nómina semana     — detalle por trabajador
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError("Instala openpyxl: pip install openpyxl") from e

    if fecha is None:
        fecha = datetime.now().strftime("%Y-%m-%d")

    d    = date.fromisoformat(fecha)
    sab  = _sabado_de_semana(d)
    dom  = sab + timedelta(days=1)
    lun  = sab + timedelta(days=2)
    dias = [sab.isoformat(), dom.isoformat(), lun.isoformat()]

    h_font, h_fill, h_align, gold_font, gold_fill, green_font, red_font = _estilos_base()

    # Cargar datos
    with _conexion() as conn:
        with conn.cursor() as cur:
            # Ventas por dia
            v_por_dia = {}
            for fd in dias:
                cur.execute(
                    """SELECT id_factura, fecha_hora, total_pagar, metodo_pago, tipo_entrega
                       FROM ventas WHERE LEFT(fecha_hora,10)=%s AND (anulada=0 OR anulada IS NULL)
                       ORDER BY fecha_hora""", (fd,),
                )
                v_por_dia[fd] = [dict(r) for r in cur.fetchall()]

            # Insumos por dia
            ins_por_dia = {}
            for fd in dias:
                cur.execute(
                    """SELECT cd.nombre_insumo, cd.cantidad, cd.unidad, cd.valor_unitario, cd.subtotal, c.fecha_hora
                       FROM compras_detalle cd JOIN compras c ON cd.compra_id=c.id
                       WHERE c.fecha=%s ORDER BY c.fecha_hora""", (fd,),
                )
                ins_por_dia[fd] = [dict(r) for r in cur.fetchall()]

            # Nómina de la semana
            cur.execute(
                "SELECT id, total, estado, fecha_pago FROM nomina_semana WHERE fecha_inicio=%s LIMIT 1",
                (sab.isoformat(),),
            )
            ns = cur.fetchone()
            detalle_nom = []
            if ns:
                cur.execute(
                    """SELECT nombre_trabajador, rol, tarifa_dia, trabajo_sabado, trabajo_domingo,
                              trabajo_lunes, dias_normales, dias_festivos, total_trabajador
                       FROM nomina_detalle WHERE nomina_id=%s ORDER BY nombre_trabajador""",
                    (ns["id"],),
                )
                detalle_nom = [dict(r) for r in cur.fetchall()]

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Resumen semana"
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = f"Álvarez Fast Food — Semana {sab.isoformat()} / {dom.isoformat()} / {lun.isoformat()}"
    ws1["A1"].font  = gold_font

    labels_dias = ["Sábado", "Domingo", "Lunes"]
    _set_header(ws1, 3, ["Día", "Ventas", "Facturas", "Insumos"], h_font, h_fill, h_align)
    tot_v = tot_i = 0
    for ri, (fd, lbl) in enumerate(zip(dias, labels_dias), 4):
        vs = sum(r["total_pagar"] for r in v_por_dia[fd])
        ins = sum(r["subtotal"] for r in ins_por_dia[fd])
        ws1.cell(ri, 1, lbl)
        c = ws1.cell(ri, 2, vs); c.number_format = "#,##0"; c.font = gold_font
        ws1.cell(ri, 3, len(v_por_dia[fd]))
        c = ws1.cell(ri, 4, ins); c.number_format = "#,##0"; c.font = red_font
        tot_v += vs; tot_i += ins

    tot_nom = int(ns["total"]) if ns else 0
    ganancia = tot_v - tot_i - tot_nom

    for fi, (lbl, val, fnt) in enumerate([
        ("TOTAL VENTAS", tot_v, gold_font),
        ("TOTAL INSUMOS", tot_i, red_font),
        ("NÓMINA", tot_nom, red_font),
        ("GANANCIA REAL", ganancia, green_font if ganancia >= 0 else red_font),
    ], 8):
        ws1.cell(fi, 1, lbl).font = gold_font
        c = ws1.cell(fi, 2, val); c.number_format = "#,##0"; c.font = fnt
    for col, w in zip("ABCD", [18, 16, 12, 16]): ws1.column_dimensions[col].width = w

    # ── Hoja 2: Ventas ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ventas fin de semana")
    _set_header(ws2, 1, ["Fecha", "Factura", "Hora", "Método", "Entrega", "Total"], h_font, h_fill, h_align)
    for fd, lbl in zip(dias, labels_dias):
        for v in v_por_dia[fd]:
            fh = str(v["fecha_hora"]); hora = fh[11:16] if len(fh) > 11 else ""
            ws2.append([lbl, v["id_factura"], hora, v["metodo_pago"], v["tipo_entrega"], int(v["total_pagar"])])
            ws2.cell(ws2.max_row, 6).number_format = "#,##0"
    for col, w in zip("ABCDEF", [12, 18, 10, 14, 12, 14]): ws2.column_dimensions[col].width = w

    # ── Hoja 3: Insumos ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Insumos semana")
    _set_header(ws3, 1, ["Fecha", "Hora", "Insumo", "Cantidad", "Unidad", "Valor unit.", "Subtotal"],
                h_font, h_fill, h_align)
    tot_ins = 0
    for fd, lbl in zip(dias, labels_dias):
        for r in ins_por_dia[fd]:
            fh = str(r["fecha_hora"]); hora = fh[11:16] if len(fh) > 11 else ""
            ws3.append([lbl, hora, r["nombre_insumo"], float(r["cantidad"]), r["unidad"],
                        int(r["valor_unitario"]), int(r["subtotal"])])
            ws3.cell(ws3.max_row, 6).number_format = "#,##0"
            ws3.cell(ws3.max_row, 7).number_format = "#,##0"
            tot_ins += int(r["subtotal"])
    fr = ws3.max_row + 2
    ws3.cell(fr, 1, "TOTAL INSUMOS").font = gold_font
    c = ws3.cell(fr, 7, tot_ins); c.number_format = "#,##0"; c.font = red_font
    for col, w in zip("ABCDEFG", [12, 10, 28, 12, 10, 14, 14]): ws3.column_dimensions[col].width = w

    # ── Hoja 4: Nómina ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Nomina semana")
    if ns:
        ws4["A1"].value = f"Nómina semana {sab.isoformat()} — Estado: {ns['estado']}"
        ws4["A1"].font  = gold_font
    _set_header(ws4, 3, ["Trabajador", "Rol", "Tarifa/día", "Sáb", "Dom", "Lun",
                          "Días norm.", "Días fest.", "Total"], h_font, h_fill, h_align)
    tot_nom_det = 0
    for r in detalle_nom:
        ws4.append([r["nombre_trabajador"], r["rol"], int(r["tarifa_dia"]),
                    int(r["trabajo_sabado"]), int(r["trabajo_domingo"]), int(r["trabajo_lunes"]),
                    int(r["dias_normales"]), int(r["dias_festivos"]), int(r["total_trabajador"])])
        ws4.cell(ws4.max_row, 3).number_format = "#,##0"
        ws4.cell(ws4.max_row, 9).number_format = "#,##0"
        tot_nom_det += int(r["total_trabajador"])
    fr4 = ws4.max_row + 2
    ws4.cell(fr4, 1, "TOTAL NÓMINA").font = gold_font
    c = ws4.cell(fr4, 9, tot_nom_det); c.number_format = "#,##0"; c.font = red_font
    for col, w in zip("ABCDEFGHI", [22, 14, 14, 8, 8, 8, 12, 12, 14]): ws4.column_dimensions[col].width = w

    if output is not None:
        wb.save(output); return output
    dest = DATA_DIR / f"alvarez_semana_{sab.isoformat()}.xlsx"
    wb.save(dest); return dest


# ════════════════════════════════════════════════════════════════════════════
# MES — 3 hojas
# ════════════════════════════════════════════════════════════════════════════

def exportar_mes(anio: int | None = None, mes: int | None = None, output=None):
    """
    Excel mensual (3 hojas):
      1. Resumen mensual   — totales + ganancia + ventas por día
      2. Ventas del mes    — una fila por factura
      3. Nómina del mes    — semanas pagadas del mes
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError("Instala openpyxl: pip install openpyxl") from e

    hoy = datetime.now()
    if anio is None: anio = hoy.year
    if mes  is None: mes  = hoy.month

    _, dias_mes = calendar.monthrange(anio, mes)
    f_ini = date(anio, mes, 1).isoformat()
    f_fin = date(anio, mes, dias_mes).isoformat()
    nom_mes = MESES_ES[mes - 1]

    h_font, h_fill, h_align, gold_font, gold_fill, green_font, red_font = _estilos_base()

    with _conexion() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT COALESCE(SUM(total_pagar),0) AS tv, COUNT(*) AS tf,
                          COALESCE(SUM(CASE WHEN metodo_pago='Efectivo' THEN total_pagar ELSE 0 END),0) AS ef,
                          COALESCE(SUM(CASE WHEN metodo_pago='Nequi'    THEN total_pagar ELSE 0 END),0) AS nq
                   FROM ventas
                   WHERE LEFT(fecha_hora,10) BETWEEN %s AND %s AND (anulada=0 OR anulada IS NULL)""",
                (f_ini, f_fin),
            )
            rv = cur.fetchone()
            cur.execute(
                """SELECT LEFT(fecha_hora,10) AS fecha, COALESCE(SUM(total_pagar),0) AS total, COUNT(*) AS facturas
                   FROM ventas WHERE LEFT(fecha_hora,10) BETWEEN %s AND %s AND (anulada=0 OR anulada IS NULL)
                   GROUP BY fecha ORDER BY fecha""", (f_ini, f_fin),
            )
            por_dia = [dict(r) for r in cur.fetchall()]
            cur.execute(
                "SELECT COALESCE(SUM(total),0) AS ti FROM compras WHERE fecha BETWEEN %s AND %s",
                (f_ini, f_fin),
            )
            total_ins = int(cur.fetchone()["ti"] or 0)
            cur.execute(
                """SELECT id, fecha_inicio, fecha_fin, total, estado
                   FROM nomina_semana WHERE estado='pagada' AND fecha_inicio BETWEEN %s AND %s ORDER BY fecha_inicio""",
                (f_ini, f_fin),
            )
            nominas = [dict(r) for r in cur.fetchall()]
            total_nom = sum(int(n["total"]) for n in nominas)
            cur.execute(
                """SELECT id_factura, fecha_hora, total_pagar, metodo_pago, tipo_entrega
                   FROM ventas WHERE LEFT(fecha_hora,10) BETWEEN %s AND %s AND (anulada=0 OR anulada IS NULL)
                   ORDER BY fecha_hora""", (f_ini, f_fin),
            )
            ventas_mes = [dict(r) for r in cur.fetchall()]

    wb = Workbook()
    tv = int(rv["tv"] or 0)
    ganancia = tv - total_ins - total_nom

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Resumen mensual"
    ws1.merge_cells("A1:C1")
    ws1["A1"].value = f"Álvarez Fast Food — {nom_mes} {anio}"; ws1["A1"].font = gold_font

    _set_header(ws1, 3, ["Concepto", "Valor"], h_font, h_fill, h_align)
    resumen_rows = [
        ("Total ventas",    tv,          gold_font),
        ("Efectivo",        int(rv["ef"] or 0), None),
        ("Nequi",           int(rv["nq"] or 0), None),
        ("Total insumos",   total_ins,   red_font),
        ("Total nómina",    total_nom,   red_font),
        ("Ganancia real",   ganancia,    green_font if ganancia >= 0 else red_font),
    ]
    for ri, (lbl, val, fnt) in enumerate(resumen_rows, 4):
        ws1.cell(ri, 1, lbl)
        c = ws1.cell(ri, 2, val); c.number_format = "#,##0"
        if fnt: c.font = fnt

    ws1.cell(12, 1, f"Total facturas: {int(rv['tf'] or 0)}")
    ws1.cell(13, 1, f"Semanas de nómina: {len(nominas)}")

    row_pd = 15
    ws1.cell(row_pd - 1, 1, "VENTAS POR DÍA").font = gold_font
    _set_header(ws1, row_pd, ["Fecha", "Total", "Facturas"], h_font, h_fill, h_align)
    for r in por_dia:
        row_pd += 1
        ws1.cell(row_pd, 1, str(r["fecha"]))
        c = ws1.cell(row_pd, 2, int(r["total"])); c.number_format = "#,##0"
        ws1.cell(row_pd, 3, int(r["facturas"]))
    for col, w in zip("ABC", [18, 16, 12]): ws1.column_dimensions[col].width = w

    # ── Hoja 2: Ventas ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ventas del mes")
    _set_header(ws2, 1, ["Factura", "Fecha", "Hora", "Método", "Entrega", "Total"], h_font, h_fill, h_align)
    for v in ventas_mes:
        fh = str(v["fecha_hora"]); fecha_v = fh[:10]; hora_v = fh[11:16] if len(fh) > 11 else ""
        ws2.append([v["id_factura"], fecha_v, hora_v, v["metodo_pago"], v["tipo_entrega"], int(v["total_pagar"])])
        ws2.cell(ws2.max_row, 6).number_format = "#,##0"
    fr2 = ws2.max_row + 2
    ws2.cell(fr2, 5, "TOTAL").font = gold_font
    c = ws2.cell(fr2, 6, tv); c.number_format = "#,##0"; c.font = gold_font
    for col, w in zip("ABCDEF", [18, 12, 10, 14, 12, 14]): ws2.column_dimensions[col].width = w

    # ── Hoja 3: Nómina ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Nomina del mes")
    _set_header(ws3, 1, ["Semana inicio", "Semana fin", "Total", "Estado"], h_font, h_fill, h_align)
    for n in nominas:
        ws3.append([str(n["fecha_inicio"]), str(n["fecha_fin"]), int(n["total"]), n["estado"]])
        ws3.cell(ws3.max_row, 3).number_format = "#,##0"
    fr3 = ws3.max_row + 2
    ws3.cell(fr3, 2, "TOTAL NÓMINA").font = gold_font
    c = ws3.cell(fr3, 3, total_nom); c.number_format = "#,##0"; c.font = red_font
    for col, w in zip("ABCD", [16, 14, 16, 12]): ws3.column_dimensions[col].width = w

    if output is not None:
        wb.save(output); return output
    dest = DATA_DIR / f"alvarez_{anio}_{mes:02d}.xlsx"
    wb.save(dest); return dest


# ════════════════════════════════════════════════════════════════════════════
# TRABAJADOR — 2 hojas
# ════════════════════════════════════════════════════════════════════════════

def exportar_trabajador(trabajador_id: int, desde: str, hasta: str, output=None):
    """
    Excel por trabajador (2 hojas):
      1. Resumen trabajador — info + totales
      2. Semanas trabajadas — tabla por semana
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError("Instala openpyxl: pip install openpyxl") from e

    h_font, h_fill, h_align, gold_font, gold_fill, green_font, red_font = _estilos_base()

    with _conexion() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, nombre, rol, tarifa_dia FROM trabajadores WHERE id=%s", (trabajador_id,))
            t = cur.fetchone()
            if not t:
                raise ValueError(f"Trabajador {trabajador_id} no encontrado")
            cur.execute(
                """SELECT nd.trabajo_sabado, nd.trabajo_domingo, nd.trabajo_lunes,
                          nd.dias_normales, nd.dias_festivos, nd.total_trabajador,
                          ns.fecha_inicio, ns.fecha_fin, ns.estado
                   FROM nomina_detalle nd
                   JOIN nomina_semana ns ON ns.id=nd.nomina_id
                   WHERE nd.trabajador_id=%s AND ns.fecha_inicio BETWEEN %s AND %s
                   ORDER BY ns.fecha_inicio""",
                (trabajador_id, desde, hasta),
            )
            semanas = [dict(r) for r in cur.fetchall()]

    total_ganado = sum(int(s["total_trabajador"]) for s in semanas)
    dias_sab = sum(int(s["trabajo_sabado"])  for s in semanas)
    dias_dom = sum(int(s["trabajo_domingo"]) for s in semanas)
    dias_lun = sum(int(s["trabajo_lunes"])   for s in semanas)
    dias_fest = sum(int(s["dias_festivos"])  for s in semanas)

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Resumen trabajador"
    ws1["A1"].value = f"Álvarez Fast Food — {t['nombre']}"; ws1["A1"].font = gold_font
    ws1["A2"].value = f"Rol: {t['rol'] or '—'}  |  Tarifa: ${int(t['tarifa_dia']):,}/día".replace(",", ".")

    info_rows = [
        ("Período", f"{desde} → {hasta}"),
        ("Total ganado",   f"${total_ganado:,}".replace(",", ".")),
        ("Semanas",        str(len(semanas))),
        ("Días sábado",    str(dias_sab)),
        ("Días domingo",   str(dias_dom)),
        ("Días lunes",     str(dias_lun)),
        ("Días festivos",  str(dias_fest)),
    ]
    for ri, (lbl, val) in enumerate(info_rows, 4):
        ws1.cell(ri, 1, lbl)
        c = ws1.cell(ri, 2, val)
        if lbl == "Total ganado": c.font = green_font
    for col, w in zip("AB", [18, 24]): ws1.column_dimensions[col].width = w

    # ── Hoja 2: Semanas ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Semanas trabajadas")
    _set_header(ws2, 1, ["Semana inicio", "Semana fin", "Sáb", "Dom", "Lun",
                          "Días norm.", "Días fest.", "Estado", "Total"],
                h_font, h_fill, h_align)
    for s in semanas:
        ws2.append([str(s["fecha_inicio"]), str(s["fecha_fin"]),
                    int(s["trabajo_sabado"]), int(s["trabajo_domingo"]), int(s["trabajo_lunes"]),
                    int(s["dias_normales"]), int(s["dias_festivos"]), s["estado"],
                    int(s["total_trabajador"])])
        ws2.cell(ws2.max_row, 9).number_format = "#,##0"
    fr = ws2.max_row + 2
    ws2.cell(fr, 8, "TOTAL").font = gold_font
    c = ws2.cell(fr, 9, total_ganado); c.number_format = "#,##0"; c.font = green_font
    for col, w in zip("ABCDEFGHI", [16, 14, 8, 8, 8, 12, 12, 12, 14]): ws2.column_dimensions[col].width = w

    if output is not None:
        wb.save(output); return output
    dest = DATA_DIR / f"alvarez_trabajador_{trabajador_id}_{desde}_{hasta}.xlsx"
    wb.save(dest); return dest


# ════════════════════════════════════════════════════════════════════════════
# GENERAL — 6 hojas
# ════════════════════════════════════════════════════════════════════════════

def exportar_general(desde: str, hasta: str, output=None):
    """
    Excel general (6 hojas):
      1. Resumen financiero — totales del período
      2. Ventas día a día   — resumen por día
      3. Insumos día a día  — compras agrupadas por día
      4. Nómina semanas     — semanas pagadas
      5. Top productos      — ranking de los más vendidos
      6. Detalle ventas     — todas las facturas del período (max 500)
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError("Instala openpyxl: pip install openpyxl") from e

    h_font, h_fill, h_align, gold_font, gold_fill, green_font, red_font = _estilos_base()

    with _conexion() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT COALESCE(SUM(total_pagar),0) AS tv, COUNT(*) AS tf,
                          COALESCE(SUM(CASE WHEN metodo_pago='Efectivo' THEN total_pagar ELSE 0 END),0) AS ef,
                          COALESCE(SUM(CASE WHEN metodo_pago='Nequi'    THEN total_pagar ELSE 0 END),0) AS nq,
                          COALESCE(SUM(CASE WHEN metodo_pago NOT IN ('Efectivo','Nequi') THEN total_pagar ELSE 0 END),0) AS otros
                   FROM ventas
                   WHERE LEFT(fecha_hora,10) BETWEEN %s AND %s AND (anulada=0 OR anulada IS NULL)""",
                (desde, hasta),
            )
            rv = cur.fetchone()
            cur.execute(
                """SELECT LEFT(fecha_hora,10) AS fecha, COALESCE(SUM(total_pagar),0) AS total, COUNT(*) AS facturas
                   FROM ventas WHERE LEFT(fecha_hora,10) BETWEEN %s AND %s AND (anulada=0 OR anulada IS NULL)
                   GROUP BY fecha ORDER BY fecha""", (desde, hasta),
            )
            ventas_dia = [dict(r) for r in cur.fetchall()]
            cur.execute(
                """SELECT c.fecha, COALESCE(SUM(cd.subtotal),0) AS total, COUNT(DISTINCT c.id) AS compras
                   FROM compras c JOIN compras_detalle cd ON cd.compra_id=c.id
                   WHERE c.fecha BETWEEN %s AND %s GROUP BY c.fecha ORDER BY c.fecha""",
                (desde, hasta),
            )
            ins_dia = [dict(r) for r in cur.fetchall()]
            cur.execute(
                "SELECT COALESCE(SUM(total),0) AS ti FROM compras WHERE fecha BETWEEN %s AND %s",
                (desde, hasta),
            )
            total_ins = int(cur.fetchone()["ti"] or 0)
            cur.execute(
                """SELECT id, fecha_inicio, fecha_fin, total, estado
                   FROM nomina_semana WHERE estado='pagada' AND fecha_inicio BETWEEN %s AND %s ORDER BY fecha_inicio""",
                (desde, hasta),
            )
            nominas = [dict(r) for r in cur.fetchall()]
            total_nom = sum(int(n["total"]) for n in nominas)
            cur.execute(
                """SELECT lv.producto_nombre, SUM(lv.cantidad) AS cant, SUM(lv.cantidad*lv.precio_unitario) AS total
                   FROM lineas_venta lv JOIN ventas v ON v.id_factura=lv.id_factura
                   WHERE LEFT(v.fecha_hora,10) BETWEEN %s AND %s AND (v.anulada=0 OR v.anulada IS NULL)
                   GROUP BY lv.producto_nombre ORDER BY total DESC LIMIT 20""",
                (desde, hasta),
            )
            top_p = [dict(r) for r in cur.fetchall()]
            cur.execute(
                """SELECT id_factura, fecha_hora, total_pagar, metodo_pago, tipo_entrega
                   FROM ventas WHERE LEFT(fecha_hora,10) BETWEEN %s AND %s AND (anulada=0 OR anulada IS NULL)
                   ORDER BY fecha_hora LIMIT 500""",
                (desde, hasta),
            )
            detalle_v = [dict(r) for r in cur.fetchall()]

    tv = int(rv["tv"] or 0)
    ganancia = tv - total_ins - total_nom

    wb = Workbook()

    # ── Hoja 1: Resumen financiero ────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Resumen financiero"
    ws1.merge_cells("A1:C1")
    ws1["A1"].value = f"Álvarez Fast Food — General {desde} → {hasta}"; ws1["A1"].font = gold_font
    _set_header(ws1, 3, ["Concepto", "Valor"], h_font, h_fill, h_align)
    rows_res = [
        ("Total ventas",       tv,                              gold_font),
        ("  Efectivo",         int(rv["ef"] or 0),              None),
        ("  Nequi",            int(rv["nq"] or 0),              None),
        ("  Otros",            int(rv["otros"] or 0),           None),
        ("Total facturas",     int(rv["tf"] or 0),              None),
        ("Total insumos",      total_ins,                       red_font),
        ("Total nómina",       total_nom,                       red_font),
        (f"  Semanas ({len(nominas)})", "",                     None),
        ("Ganancia real",      ganancia,                        green_font if ganancia >= 0 else red_font),
    ]
    for ri, (lbl, val, fnt) in enumerate(rows_res, 4):
        ws1.cell(ri, 1, lbl)
        if val != "":
            c = ws1.cell(ri, 2, val); c.number_format = "#,##0"
            if fnt: c.font = fnt
    for col, w in zip("AB", [22, 18]): ws1.column_dimensions[col].width = w

    # ── Hoja 2: Ventas día a día ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Ventas dia a dia")
    _set_header(ws2, 1, ["Fecha", "Total ventas", "Facturas"], h_font, h_fill, h_align)
    tot_v_dia = 0
    for d in ventas_dia:
        ws2.append([str(d["fecha"]), int(d["total"]), int(d["facturas"])])
        ws2.cell(ws2.max_row, 2).number_format = "#,##0"
        tot_v_dia += int(d["total"])
    fr2 = ws2.max_row + 2
    ws2.cell(fr2, 1, "TOTAL").font = gold_font
    c = ws2.cell(fr2, 2, tot_v_dia); c.number_format = "#,##0"; c.font = gold_font
    for col, w in zip("ABC", [14, 16, 12]): ws2.column_dimensions[col].width = w

    # ── Hoja 3: Insumos día a día ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Insumos dia a dia")
    _set_header(ws3, 1, ["Fecha", "Total insumos", "Compras"], h_font, h_fill, h_align)
    tot_i_dia = 0
    for d in ins_dia:
        ws3.append([str(d["fecha"]), int(d["total"]), int(d["compras"])])
        ws3.cell(ws3.max_row, 2).number_format = "#,##0"
        tot_i_dia += int(d["total"])
    fr3 = ws3.max_row + 2
    ws3.cell(fr3, 1, "TOTAL").font = gold_font
    c = ws3.cell(fr3, 2, tot_i_dia); c.number_format = "#,##0"; c.font = red_font
    for col, w in zip("ABC", [14, 16, 12]): ws3.column_dimensions[col].width = w

    # ── Hoja 4: Nómina semanas ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Nomina semanas")
    _set_header(ws4, 1, ["Semana inicio", "Semana fin", "Total", "Estado"], h_font, h_fill, h_align)
    for n in nominas:
        ws4.append([str(n["fecha_inicio"]), str(n["fecha_fin"]), int(n["total"]), n["estado"]])
        ws4.cell(ws4.max_row, 3).number_format = "#,##0"
    fr4 = ws4.max_row + 2
    ws4.cell(fr4, 2, "TOTAL NÓMINA").font = gold_font
    c = ws4.cell(fr4, 3, total_nom); c.number_format = "#,##0"; c.font = red_font
    for col, w in zip("ABCD", [16, 14, 16, 12]): ws4.column_dimensions[col].width = w

    # ── Hoja 5: Top productos ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Top productos")
    _set_header(ws5, 1, ["#", "Producto", "Cantidad vendida", "Total"], h_font, h_fill, h_align)
    for i, p in enumerate(top_p, 1):
        ws5.append([i, p["producto_nombre"], int(p["cant"]), int(p["total"])])
        ws5.cell(ws5.max_row, 4).number_format = "#,##0"
    for col, w in zip("ABCD", [6, 36, 18, 16]): ws5.column_dimensions[col].width = w

    # ── Hoja 6: Detalle ventas ────────────────────────────────────────────────
    ws6 = wb.create_sheet("Detalle ventas")
    _set_header(ws6, 1, ["Factura", "Fecha", "Hora", "Método", "Entrega", "Total"],
                h_font, h_fill, h_align)
    for v in detalle_v:
        fh = str(v["fecha_hora"]); fecha_v = fh[:10]; hora_v = fh[11:16] if len(fh) > 11 else ""
        ws6.append([v["id_factura"], fecha_v, hora_v, v["metodo_pago"], v["tipo_entrega"], int(v["total_pagar"])])
        ws6.cell(ws6.max_row, 6).number_format = "#,##0"
    fr6 = ws6.max_row + 2
    ws6.cell(fr6, 5, "TOTAL").font = gold_font
    c = ws6.cell(fr6, 6, tv); c.number_format = "#,##0"; c.font = gold_font
    for col, w in zip("ABCDEF", [18, 12, 10, 14, 12, 14]): ws6.column_dimensions[col].width = w

    if output is not None:
        wb.save(output); return output
    dest = DATA_DIR / f"alvarez_general_{desde}_{hasta}.xlsx"
    wb.save(dest); return dest
